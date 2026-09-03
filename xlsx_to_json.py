#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
将本地 birthday.xlsx 转换为 birthday.json，作为 Cloudflare Worker 的唯一数据源。

- 只读取本地 xlsx，不依赖 pandas（使用 openpyxl），兼容日期单元格与字符串。
- birthday.xlsx 本身不会上传到任何线上环境；线上程序只解析生成的 json。
- 输出: worker/src/birthday.json

用法:
    python xlsx_to_json.py
    python xlsx_to_json.py 输入.xlsx 输出.json
"""

import os
import sys
import json
import datetime
import openpyxl

# 列名解析：兼容中英文表头
COLUMN_ALIASES = {
    "name": ["name", "姓名", "名字"],
    "type": ["type", "类型", "历"],
    "birthday": ["birthday", "生日", "阳历生日"],
    "lunarbirthday": ["lunarbirthday", "阴历生日", "农历生日", "lunar"],
    "note": ["备注", "note", "备注1"],
}


def _header_index(header_cells, aliases):
    """在表头行中按别名（不区分大小写）查找列索引，返回 0-based，找不到返回 None。"""
    norm = {}
    for idx, raw in enumerate(header_cells):
        if raw is None:
            continue
        norm[str(raw).strip().lower()] = idx
    for alias in aliases:
        if alias.lower() in norm:
            return norm[alias.lower()]
    # 模糊匹配：别名是表头子串
    for idx, raw in enumerate(header_cells):
        if raw is None:
            continue
        low = str(raw).strip().lower()
        for alias in aliases:
            if alias.lower() in low:
                return idx
    return None


def _to_date(value):
    """把单元格值规整为 (year, month, day)；无法解析返回 None。"""
    if value is None:
        return None
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.year, value.month, value.day
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"):
            try:
                d = datetime.datetime.strptime(s, fmt)
                return d.year, d.month, d.day
            except ValueError:
                continue
    return None


def _is_solar(type_value):
    if not type_value:
        return False
    return "阳" in str(type_value)


def convert(xlsx_path, json_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Sheet1"]

    header = [c.value for c in ws[1]]
    idx = {k: _header_index(header, v) for k, v in COLUMN_ALIASES.items()}

    people = []
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(row=r, column=c + 1).value for c in range(len(header))]
        name = row[idx["name"]] if idx["name"] is not None else None
        if name is None or str(name).strip() == "":
            continue
        name = str(name).strip()

        type_value = row[idx["type"]] if idx["type"] is not None else None
        solar = _to_date(row[idx["birthday"]]) if idx["birthday"] is not None else None
        lunar = (
            _to_date(row[idx["lunarbirthday"]])
            if idx["lunarbirthday"] is not None
            else None
        )
        note = row[idx["note"]] if idx["note"] is not None else None
        if note is not None:
            note = str(note).strip()
            if note == "":
                note = None

        # 阳历类型为基准；阴历类型用 lunar 列计算，但年龄仍以阳历年份为准
        if solar is None:
            # 退而求其次：用 lunar 年份补阳历年份，避免年龄算错
            solar = lunar

        person = {
            "name": name,
            "type": "阳历" if _is_solar(type_value) else "阴历",
            "solar": {"year": solar[0], "month": solar[1], "day": solar[2]} if solar else None,
            "lunar": {"year": lunar[0], "month": lunar[1], "day": lunar[2]} if lunar else None,
        }
        if note:
            person["note"] = note
        people.append(person)

    payload = {
        "updatedAt": datetime.date.today().isoformat(),
        "source": os.path.basename(xlsx_path),
        "people": people,
    }

    os.makedirs(os.path.dirname(json_path), exist_ok=True)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"已转换 {len(people)} 条记录 -> {json_path}")
    return payload


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    in_path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(script_dir, "birthday.xlsx")
    out_path = (
        sys.argv[2]
        if len(sys.argv) > 2
        else os.path.join(script_dir, "worker", "src", "birthday.json")
    )
    if not os.path.exists(in_path):
        print(f"找不到输入文件: {in_path}")
        sys.exit(1)
    convert(in_path, out_path)


if __name__ == "__main__":
    main()
